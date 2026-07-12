import os
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def format_params(params):
    if not params:
        return ""
    if isinstance(params, str):
        try:
            params = json.loads(params.replace("'", '"'))
        except:
            return params
    # Vertikale Liste der Parameter
    return "\n".join([f"{k}: {v}" for k, v in params.items()])

def main():
    # Ermittle den Projekt-Wurzelordner basierend auf dem Skript-Speicherort
    script_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.abspath(os.path.join(script_dir, "..", ".."))
    
    json_path = os.path.join(project_root, 'Model/reports/final/training_metrics.json')
    excel_path = os.path.join(project_root, 'sonst/excel_export/modellvergleich_ergebnisse.xlsx')
    
    if not os.path.exists(json_path):
        print(f"Fehler: {json_path} existiert nicht.")
        return
        
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
        
    # Sicherstellen, dass das Export-Verzeichnis existiert
    os.makedirs(os.path.dirname(excel_path), exist_ok=True)
        
    wb = openpyxl.Workbook()
    # Standardblatt entfernen
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Schriftarten, Füllungen und Rahmen
    font_title = Font(name='Calibri', size=14, bold=True, color='1F497D')
    font_section = Font(name='Calibri', size=12, bold=True)
    font_header = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    font_data = Font(name='Calibri', size=11)
    font_best_text = Font(name='Calibri', size=10, italic=True, color='3F3F3F')
    
    fill_header = PatternFill(start_color='1F497D', end_color='1F497D', fill_type='solid')
    fill_winner = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid') # Soft green
    
    border_thin = Side(style='thin', color='D9D9D9')
    border_data = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    
    headers = ["Modell", "R²-Score", "RMSE (%)", "MSE (%)", "MAE (%)", "Beste Hyperparameter"]
    
    # ---------------- 1. SHEET: QUICK MODUS ----------------
    ws_quick = wb.create_sheet(title="Quick Modus")
    ws_quick.views.sheetView[0].showGridLines = True
    
    ws_quick.append([])
    ws_quick.cell(row=2, column=1, value="Ergebnisse: Quick (Tagesplaner)").font = font_title
    ws_quick.append([])
    
    ws_quick.append(headers)
    
    # Style header
    for col_idx in range(1, len(headers) + 1):
        cell = ws_quick.cell(row=4, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal='center' if col_idx > 1 and col_idx < 6 else 'left', vertical='center')
        
    quick_models = data.get('Quick', {})
    best_q_name = ""
    best_q_r2 = -9999.0
    for name, m in quick_models.items():
        r2 = m.get('R2', 0.0)
        if r2 > best_q_r2:
            best_q_r2 = r2
            best_q_name = name
            
    row_num = 5
    for name, m in quick_models.items():
        row_data = [
            name,
            m.get('R2', 0.0),
            m.get('RMSE', 0.0),
            m.get('MSE', 0.0),
            m.get('MAE', 0.0),
            format_params(m.get('best_params', {}))
        ]
        ws_quick.append(row_data)
        
        is_winner = (name == best_q_name)
        for col_idx in range(1, len(headers) + 1):
            cell = ws_quick.cell(row=row_num, column=col_idx)
            cell.font = font_data
            cell.border = border_data
            if is_winner:
                cell.fill = fill_winner
            
            # Ausrichtung immer oben bündig (top-aligned) mit automatischem Zeilenumbruch
            h_align = 'left'
            if col_idx in [2, 3, 4, 5]:
                h_align = 'center'
            cell.alignment = Alignment(horizontal=h_align, vertical='top', wrap_text=True)
            
            # Zahlenformate
            if col_idx == 2:
                cell.number_format = '0.0000'
            elif col_idx in [3, 4, 5]:
                cell.number_format = '0.00'
        row_num += 1
        
    ws_quick.append([])
    best_cell = ws_quick.cell(row=row_num+1, column=1, value=f"Bestes Modell für Quick: {best_q_name} (R² = {best_q_r2:.4f})")
    best_cell.font = font_best_text
    
    # Spaltenbreiten mit großzügigen Mindestbreiten (Padding) für PowerPoint-Kompatibilität
    column_widths = {
        1: 18,  # Modell
        2: 16,  # R²-Score
        3: 16,  # RMSE (%)
        4: 16,  # MSE (%)
        5: 16,  # MAE (%)
        6: 50   # Beste Hyperparameter (bietet genug Platz für vertikale Liste)
    }
    for col in ws_quick.columns:
        max_len = 0
        for cell in col:
            val = str(cell.value or '')
            # Überspringe Titel-, Best-Modell- und Ergebnistexte bei der Längenberechnung,
            # da diese sonst die Spaltenbreite von Spalte 1 unnötig aufblähen.
            if cell.row == 2 or "Bestes Modell" in val or "Ergebnisse" in val:
                continue
            # Bei Zeilenumbrüchen die Länge der längsten Zeile nehmen
            lines = val.split('\n')
            for line in lines:
                if len(line) > max_len:
                    max_len = len(line)
        col_letter = get_column_letter(col[0].column)
        col_idx = col[0].column
        min_w = column_widths.get(col_idx, 12)
        ws_quick.column_dimensions[col_letter].width = max(max_len + 6, min_w)
        
        
    # ---------------- 2. SHEET: DETAILED MODUS ----------------
    ws_det = wb.create_sheet(title="Detailed Modus")
    ws_det.views.sheetView[0].showGridLines = True
    
    ws_det.append([])
    ws_det.cell(row=2, column=1, value="Ergebnisse: Detailed (Live-Tracker)").font = font_title
    ws_det.append([])
    
    det_data = data.get('Detailed', {})
    horizons_sorted = sorted([int(k) for k in det_data.keys()])
    
    summary_data = []
    
    current_row = 4
    for mins in horizons_sorted:
        h_key = str(mins)
        h_data = det_data[h_key]
        
        ws_det.cell(row=current_row, column=1, value=f"Zeithorizont: {mins} Minuten").font = font_section
        current_row += 1
        
        ws_det.append(headers)
        # Style header
        for col_idx in range(1, len(headers) + 1):
            cell = ws_det.cell(row=current_row, column=col_idx)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal='center' if col_idx > 1 and col_idx < 6 else 'left', vertical='center')
        current_row += 1
        
        # Find best model for this horizon
        best_h_name = ""
        best_h_r2 = -9999.0
        for name, m in h_data.items():
            r2 = m.get('R2', 0.0)
            if r2 > best_h_r2:
                best_h_r2 = r2
                best_h_name = name
                
        best_m = h_data[best_h_name]
        summary_data.append({
            'mins': mins,
            'name': best_h_name,
            'R2': best_m.get('R2', 0.0),
            'RMSE': best_m.get('RMSE', 0.0),
            'MSE': best_m.get('MSE', 0.0),
            'MAE': best_m.get('MAE', 0.0),
            'best_params': best_m.get('best_params', {})
        })
                
        for name, m in h_data.items():
            row_data = [
                name,
                m.get('R2', 0.0),
                m.get('RMSE', 0.0),
                m.get('MSE', 0.0),
                m.get('MAE', 0.0),
                format_params(m.get('best_params', {}))
            ]
            ws_det.append(row_data)
            
            is_winner = (name == best_h_name)
            for col_idx in range(1, len(headers) + 1):
                cell = ws_det.cell(row=current_row, column=col_idx)
                cell.font = font_data
                cell.border = border_data
                if is_winner:
                    cell.fill = fill_winner
                
                # Ausrichtung immer oben bündig (top-aligned) mit automatischem Zeilenumbruch
                h_align = 'left'
                if col_idx in [2, 3, 4, 5]:
                    h_align = 'center'
                cell.alignment = Alignment(horizontal=h_align, vertical='top', wrap_text=True)
                
                # Number formatting
                if col_idx == 2:
                    cell.number_format = '0.0000'
                elif col_idx in [3, 4, 5]:
                    cell.number_format = '0.00'
            current_row += 1
            
        ws_det.append([])
        best_cell = ws_det.cell(row=current_row+1, column=1, value=f"Bestes Modell für {mins}m: {best_h_name} (R² = {best_h_r2:.4f})")
        best_cell.font = font_best_text
        current_row += 3
        ws_det.append([])
        ws_det.append([])
        
    # ---------------- ZUSAMMENFASSUNGSTABELLE DETAILED MODUS ----------------
    current_row = ws_det.max_row + 3 # 3 Leerzeilen Abstand
    
    ws_det.cell(row=current_row, column=1, value="Zusammenfassung: Beste Modelle im Vergleich").font = font_title
    
    # 1 Leerzeile Abstand unter dem Titel
    ws_det.append([])
    
    # Header für die Zusammenfassungstabelle
    ws_det.append(headers)
    current_row = ws_det.max_row
    for col_idx in range(1, len(headers) + 1):
        cell = ws_det.cell(row=current_row, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal='center' if col_idx > 1 and col_idx < 6 else 'left', vertical='center')
        
    # Zeilen für jedes beste Modell schreiben
    for item in summary_data:
        row_data = [
            f"{item['name']} ({item['mins']}m)",
            item['R2'],
            item['RMSE'],
            item['MSE'],
            item['MAE'],
            format_params(item['best_params'])
        ]
        ws_det.append(row_data)
        current_row = ws_det.max_row
        
        # Jede Zeile in dieser Tabelle ist ein Siegermodell, daher färben wir sie mit fill_winner
        for col_idx in range(1, len(headers) + 1):
            cell = ws_det.cell(row=current_row, column=col_idx)
            cell.font = font_data
            cell.border = border_data
            cell.fill = fill_winner
            
            h_align = 'left'
            if col_idx in [2, 3, 4, 5]:
                h_align = 'center'
            cell.alignment = Alignment(horizontal=h_align, vertical='top', wrap_text=True)
            
            # Formatierung
            if col_idx == 2:
                cell.number_format = '0.0000'
            elif col_idx in [3, 4, 5]:
                cell.number_format = '0.00'
        
    # Spaltenbreiten mit großzügigen Mindestbreiten (Padding) für PowerPoint-Kompatibilität.
    # Wir setzen die Werte exakt auf den vorherigen Stand zurück, damit die Spaltenbreiten
    # aller Tabellen identisch sind und sie in PowerPoint exakt übereinander passen.
    column_widths = {
        1: 18,  # Modell
        2: 16,  # R²-Score
        3: 16,  # RMSE (%)
        4: 16,  # MSE (%)
        5: 16,  # MAE (%)
        6: 50   # Beste Hyperparameter
    }
    for col in ws_det.columns:
        max_len = 0
        for cell in col:
            val = str(cell.value or '')
            # Überspringe Titel-, Zeithorizont-, Best-Modell- und Zusammenfassungs-Texte bei der Längenberechnung
            if cell.row == 2 or "Zeithorizont" in val or "Bestes Modell" in val or "Ergebnisse" in val or "Zusammenfassung" in val:
                continue
            lines = val.split('\n')
            for line in lines:
                if len(line) > max_len:
                    max_len = len(line)
        col_letter = get_column_letter(col[0].column)
        col_idx = col[0].column
        min_w = column_widths.get(col_idx, 12)
        ws_det.column_dimensions[col_letter].width = max(max_len + 6, min_w)
        
    wb.save(excel_path)
    print(f"Erfolgreich Excel-Export erstellt unter: {excel_path}")

if __name__ == '__main__':
    main()
